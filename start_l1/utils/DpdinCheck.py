import os
import sys
import logging
import json
import threading
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
current_path = os.path.dirname(os.path.realpath(__file__))
START_L1_FOLDER = os.path.dirname(current_path)
AUTO_TEST_FOLDER = os.path.dirname(START_L1_FOLDER)
DUT_CONTROL_FOLDER = os.path.dirname(AUTO_TEST_FOLDER)

try:
    from auto_test.start_l1.connections.connection import _SSH_Shell
    from auto_test.start_l1.utils.settings import download_runtime_log
except:
    sys.path.insert(0, DUT_CONTROL_FOLDER)
    from auto_test.start_l1.connections.connection import _SSH_Shell
    from auto_test.start_l1.utils.settings import download_runtime_log

START_L1_FOLDER = os.path.dirname(current_path)
class DpdinStatus():
    def __init__(self, made_ips, limit=None):
        if limit is None:
            self.limit = {'dpdin': 5, 'fb': 5, 'dpdout': 5}
        else:
            self.limit = limit
        self.made_ips = made_ips
        self.logger = logging.getLogger('main')
        self.fail_flag = False


    def DPDIN_test_singleIP(self, host, res, repeat, row, col):
        col_origin = col
        count = 0
        port = 22
        username = "root"
        command = "dapd -sq"
        ssh = _SSH_Shell()
        ssh.open(hostname=host, port=port, username=username, password=None)
        while count < repeat:
            stdout, stderr = ssh.command(command)
            self.logger.info(f"sendind command to {host}=> {command}")
            for line in stdout.split('\n'):
                self.logger.info(line)
                if line.strip().startswith('DPD IN:'):
                    dpdin_values = [float(value) for value in line.split() if value != 'DPD' and value != 'IN:'
                                    and '(' not in value and ')' not in value]
                if line.strip().startswith('DPD OUT:'):
                    dpdout_values = [float(value) for value in line.split() if value != 'DPD' and value != 'OUT:'
                                    and '(' not in value and ')' not in value]
                if line.strip().startswith('FB:'):
                    fb_values = [float(value) for value in line.split() if value != 'FB:'
                                    and '(' not in value and ')' not in value]
            try:
                print("res", dpdin_values)
                self.logger.info(f"get DPDIN values from {host}=> " + str(dpdin_values))
                self.logger.info(f"get DPDOUT values from {host}=> " + str(dpdout_values))
                self.logger.info(f"get FB values from {host}=> " + str(fb_values))
                self.ws.cell(row=row, column=col, value=time.strftime("%Y%m%d%H%M%S"))
                col += 1
                self.ws.cell(row=row, column=col, value=host)
                col += 1

                # write dpdin values
                dpdin_pass_flag = True
                for value in dpdin_values:
                    _ = self.ws.cell(row=row, column=col, value=value)
                    col += 1
                    if not (-11.3 - self.limit['dpdin'] <= value <= -11.3 + self.limit['dpdin']):
                        # print("value ot of limit!")
                        _.font = Font(color='00FF0000')
                        dpdin_pass_flag = False
                        self.fail_flag = True
                if not dpdin_pass_flag:
                    self.logger.info("DPDIN Value out of limit!")
                    res[host]["DPDIN Value out of limit"] += 1

                # write dpdout values
                dpdout_pass_flag = True
                for value in dpdout_values:
                    _ = self.ws.cell(row=row, column=col, value=value)
                    col += 1
                    if not (-14.3 - self.limit['dpdout'] <= value <= -14.3 + self.limit['dpdout']):
                        # print("value ot of limit!")
                        _.font = Font(color='00FF0000')
                        dpdin_pass_flag = False
                        self.fail_flag = True
                if not dpdout_pass_flag:
                    self.logger.info("DPDOUT Value out of limit!")
                    res[host]["DPDOUT Value out of limit"] += 1

                # write fb values
                fb_pass_flag = True
                for value in fb_values:
                    _ = self.ws.cell(row=row, column=col, value=value)
                    col += 1
                    if not (-11.3 - self.limit['fb'] <= value <= -11.3 + self.limit['fb']):
                        # print("value ot of limit!")
                        _.font = Font(color='00FF0000')
                        fb_pass_flag = False
                        self.fail_flag = True
                if not fb_pass_flag:
                    self.logger.info("FB Value out of limit!")
                    res[host]["FB Value out of limit"] += 1
                if dpdin_pass_flag and fb_pass_flag:
                    res[host]["Pass times"] += 1
            except UnboundLocalError:
                # print("No dpdin value detected!")
                self.logger.error("No dpdin value detected!")
                res[host]["No value detected"] += 1
                self.fail_flag = True
            row += 1
            col = col_origin
            count += 1
            res[host]["read counts"] += 1
            time.sleep(1)
        ssh.close()

    def DPDIN_test(self, repeat=1):
        current_path = os.path.dirname(os.path.realpath(__file__))
        try:
            with open(os.path.join(START_L1_FOLDER, "testings/PowerTestResults/PowerTestResult.json"), 'r') as f:
                res = json.load(f)
        except FileNotFoundError:
            res = {host: {"read counts": 0, "Pass times": 0, "DPDIN Value out of limit": 0, "FB Value out of limit": 0,
                          "No value detected": 0} for host in self.made_ips}

        try:
            self.wb = load_workbook(os.path.join(START_L1_FOLDER, 'testings/MainTestResults/MainResult.xlsx'))
        except FileNotFoundError:
            self.wb = Workbook()
        if "Carrier power raw data" not in self.wb.sheetnames:
            self.wb.create_sheet("Carrier power raw data")
            self.ws = self.wb["Carrier power raw data"]
            _ = self.ws.cell(row=1, column=3, value='dpdin values')
            _.font = Font(bold=True)
            _ = self.ws.cell(row=1, column=11, value='dpdout values')
            _.font = Font(bold=True)
            _ = self.ws.cell(row=1, column=19, value='fb values')
            _.font = Font(bold=True)
            for i in range(8):
                self.ws.cell(row=2, column=i+3, value=i)
                self.ws.cell(row=2, column=i+11, value=i)
                self.ws.cell(row=2, column=i+19, value=i)
        else:
            self.ws = self.wb["Carrier power raw data"]
        row = self.ws.max_row + 1
        col = 1
        print('max col', col)
        threads = []
        for host in self.made_ips:
            t = threading.Thread(target=self.DPDIN_test_singleIP, args=(host, res, repeat, row, col))
            t.start()
            threads.append(t)
            row += repeat
        for t in threads:
            t.join()
        # for host in self.made_ips:
        #     self.DPDIN_test_singleIP(host, res, repeat, row, col)
        #     row += 3
        self.wb.save(filename=os.path.join(START_L1_FOLDER, 'testings/MainTestResults/MainResult.xlsx'))
        with open(os.path.join(START_L1_FOLDER, "testings/PowerTestResults/PowerTestResult.json"), 'w') as f:
            json.dump(res, f)

        if self.fail_flag:
            download_runtime_log(path=os.path.join(START_L1_FOLDER, 'testings/MainTestResults/runtimeLog/'), log_name='power_fail')

        return not self.fail_flag


if __name__ == "__main__":
    made_ips = ['192.168.101.2', '192.168.101.3', '192.168.101.4', '192.168.101.5']

    DpdinCheckHandler = DpdinStatus(made_ips)
    DpdinCheckHandler.DPDIN_test(repeat=3)
