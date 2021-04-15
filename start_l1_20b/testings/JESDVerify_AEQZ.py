import os, sys
import time
import logging
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import json
import threading
current_path = os.path.dirname(os.path.realpath(__file__))
sys.path.insert(0, os.path.dirname(os.path.dirname(current_path)))
# add tools_dev into sys path
# not necessary if call from start.py
from tools_dev.connections.miniftp import SCP
from tools_dev.connections.connection import _SSH_Shell
from tools_dev.testings.Exceptions import *

class JesdVerify(object):
    def __init__(self, made_ips, product_type, CycleTime=0):
        self.ips = made_ips
        self.product_type = product_type
        self.check_items = self._load_json()
        self.CycleTime = CycleTime
        self.fail_flag = False

    def _download_runtime_log(self):
        print("Catching runtime log...")
        logging.info("Catching runtime log...")
        current_path = os.path.dirname(os.path.realpath(__file__))
        folder = os.path.join(current_path, 'JesdLogAndResults/runtimeLog/')
        time_current = time.strftime("%Y%m%d%H%M%S")
        beamer_ip = "192.168.101.1"

        try:
            scp = SCP(host="192.168.101.1", port=22, username="toor4nsn", password="oZPS0POrRieRtu")
            scp.connect()
            scp.write(f"/usr/bin/ccsShell.sh log -a -z runtime_{time_current}.zip")
            time.sleep(5)
            scp.download(local_path=folder+f"runtime_{time_current}.zip", remote_path=f"/ram/runtime_{time_current}.zip", mode='bin')
            scp.close()
        except:
            print("Error occur when grab log from beamer! =>" + str(sys.exc_info()[0]))

    def _load_json(self):
        if self.product_type == "AEQZ":
            json_file = "AEQZ_JESD_CHECK.json"
        elif self.product_type == "AEQV":
            json_file = "AEQV_JESD_CHECK.json"
        with open(os.path.join(current_path, "JESDTestItems", json_file), "r") as f:
            check_items = json.load(f)
        return check_items

    def _scp_connect(self, ip):

        self.scp = SCP(ip, 22)
        self.scp.connect_noPassword()
        # create ssh with no password

        local = os.path.join(current_path, 'spiafe')
        remote = '/var/tmp/spiafe'
        self.scp.upload(local, remote)

    def _load_commands(self, row, col):
        self.commands = []
        if self.product_type == "AEQZ":
            command_file = "JESDCommands_AEQZ.txt"
        elif self.product_type == "AEQV":
            command_file = "JESDCommands_AEQV.txt"
        with open(os.path.join(current_path, command_file), 'r') as f:
            lines = f.readlines()
        for line in lines:
            if not line.startswith('#'):
                self.commands.append(line.strip())
                command_to_write = re.search(r'/dev/spidev.*$', line)
                if command_to_write:
                    _ = self.ws.cell(column=col, row=row, value=command_to_write.group().strip())
                else:
                    _ = self.ws.cell(column=col, row=row, value=line.strip())
                row += 1
        # self.wb.save(filename='JesdVerify.xls')

    def _scp_close(self):
        self.scp.close()

    def _read_values(self, row, col):
        value_wrong = False
        for command in self.commands:
            command = command.strip()
            # logging.info(stderr.readline())
            # check with 0xfc90010c
            if command in self.check_items.keys():
                stdin, stdout, stderr = self.scp.write(command)
                out = stdout.readline().strip()
                # print("out:"+out)
                logging.debug("get respond => %s" % out)
                if not self.check_items[command]:
                    pass
                elif self.check_items[command] == "LAST2BIT":
                    logging.debug('get binary value => %s' % bin(int(out, 16)))
                    _ = self.ws.cell(row=row, column=col, value=bin(int(out, 16)))
                    if repr(bin(int(out, 16)))[-3:-1] == '11':
                        logging.debug('Verification result => correct')
                    else:
                        value_wrong = True
                        logging.debug('Verification result => wrong')
                        _.font = Font(color='00FF0000')
                elif self.check_items[command] == "CHECKLINK":
                    _ = self.ws.cell(row=row, column=col, value=out)
                    if out in ['0x00002860', '0x00000060']:
                        logging.debug('Verification result => correct')
                    else:
                        value_wrong = True
                        logging.debug('Verification result => wrong')
                        _.font = Font(color='00FF0000')
                else:
                    _ = self.ws.cell(row=row, column=col, value=out[-2:])
                    if out.split()[-1] == self.check_items[command]:
                        logging.debug('Verification result => correct')
                    else:
                        value_wrong = True
                        logging.debug('Verification result => wrong')
                        _.font = Font(color='00FF0000')
                row += 1
            else:
                logging.info("Invalid command: " + command)
        return value_wrong


    def JesdTest_singleIP(self, REPEAT, ip, row, col):
        pass_times = 0
        fail_times = 0
        #self._load_commands(row, col)
        self._scp_connect(ip)
        for i in range(REPEAT):
            logging.info(f"reading jesd regs of {ip}")
            get_wrong_value = self._read_values(row, col+1)
            if get_wrong_value:
                self.fail_flag = True
                fail_times += 1
            else:
                pass_times += 1
            col += 1
            logging.info(f"read finished...get fail values {get_wrong_value}")
            time.sleep(3)

        logging.info('#' * 30)

        # self._scp_close()
        self.ws.cell(row=row, column=col-2, value=ip)
        self.ws.cell(row=row, column=col-1, value="Read counts:" + str(REPEAT))
        self.ws.cell(row=row, column=col, value="PASS: " + str(pass_times) + " FAIL: " + str(fail_times))
        if not self.ws.cell(row=row, column=1).value:
            self.ws.cell(row=row, column=1, value="0/0")
        total_pass = self.ws.cell(row=row, column=1).value.split("/")[0]
        total_fail = self.ws.cell(row=row, column=1).value.split("/")[1]
        total_pass = int(total_pass) + pass_times
        total_fail = int(total_fail) + fail_times
        self.ws.cell(row=row, column=1, value=str(total_pass)+"/"+str(total_fail))

    def JesdTest(self, REPEAT=1):
        current_path = os.path.dirname(os.path.realpath(__file__))
        try:
            self.wb = load_workbook(os.path.join(current_path, 'JesdLogAndResults/JesdVerify.xlsx'))
        except FileNotFoundError:
            self.wb = Workbook()
        self.ws = self.wb.active
        row = 1
        col = self.ws.max_column + 1
        self._load_commands(row, col)
        for ip in self.ips:
            logging.info("IP: "+ip)
            self.JesdTest_singleIP(REPEAT, ip, row, col)
            row += len(self.commands) + 1
        if self.fail_flag:
            self._download_runtime_log()
        self.wb.save(filename=os.path.join(current_path, 'JesdLogAndResults/JesdVerify.xlsx'))
        return



if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,  format="%(asctime)s:%(levelname)s:%(message)s")
    handler = JesdVerify(made_ips = ['192.168.101.2', '192.168.101.3', '192.168.101.4', '192.168.101.5'], product_type="AEQZ")
    handler._download_runtime_log()
