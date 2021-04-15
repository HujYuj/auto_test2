import os, sys
import time
import logging
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import json
import threading

try:
    from tools_dev.connections.miniftp import SCP
    from tools_dev.connections.connection import _SSH_Shell
    from tools_dev.testings.Exceptions import *
    from tools_dev.testings.settings import download_runtime_log
except:
    from ..connections.miniftp import SCP
    from ..connections.connection import _SSH_Shell
    from .Exceptions import *
    from .settings import download_runtime_log

# class JesdVerify(object):
#     def __init__(self, made_ips, product_type):
#         self.made_ips = made_ips
#         self.product_type = product_type
#
#     def _load_json(self):
#         """
#
#         :return: dictionary of command/value pair
#         """
#         if self.product_type == "AEQZ":
#             json_file = "AEQZ_JESD_CHECK.json"
#         elif self.product_type == "AEQV":
#             json_file = "AEQV_JESD_CHECK.json"
#         else:
#             raise ProductTypeNotFoundError
#
#         with open(os.path.join(current_path, "JESDTestItems", json_file), "r") as f:
#             check_items = json.load(f)
#         self.check_items = check_items
#         return check_items
#
#     def _load_commands(self):
#         """
#         read commands from text file
#         :return:  list
#         """
#         commands = []
#         commands_in_short = []
#         if self.product_type == "AEQZ":
#             command_file = "JESDCommands_AEQZ.txt"
#         elif self.product_type == "AEQV":
#             command_file = "JESDCommands_AEQV.txt"
#         else:
#             rasie CanNotLoadCommandsError
#         with open(os.path.join(current_path, command_file), 'r') as f:
#             lines = f.readlines()
#         for line in lines:
#             if not line.startswith('#'):
#                 commands.append(line.strip())
#                 command_in_short = re.search(r'/dev/spidev.*$', line)
#                 if command_in_short:
#                     commands_in_short.append(command_in_short.group().strip())
#                 else:
#                     commands_in_short.append(line.strip())
#         self.commands = commands
#         return commands_in_short
#
#     def _verify_jesd_single_ip(self, ip, repeat, results=None, index=None):
#         """
#         this function will return the jesd values of a single made ip with multiple reading counts
#         :return:  dict
#         """
#         scp = SCP(ip, 22)
#         scp.connect_noPassword()
#         # create ssh with no password
#         local = os.path.join(current_path, 'spiafe')
#         remote = '/var/tmp/spiafe'
#         scp.upload(local, remote)
#         print("ip", ip, "upload done")
#
#         value_read = [[] for _ in range(repeat)]
#         verify_res = []
#         for i in range(repeat):
#             for command in self.commands:
#                 if command in self.check_items.keys():
#                     stdin, stdout, stderr = scp.write(command)
#                     out = stdout.readline().strip()
#                     print("out:"+out)
#                     logging.info("get respond => %s" % out)
#                     if not self.check_items[command]:
#                         pass
#                     elif self.check_items[command] == "LAST2BIT":
#                         value_read[i].append(bin(int(out, 16)))
#                         if repr(bin(int(out, 16)))[-3:-1] == '11':
#                             verify_res.append(True)
#                         else:
#                             verify_res.append(False)
#                     elif self.check_items[command] == "CHECKLINK":
#                         value_read[i].append(out)
#                         if out in ['0x00002860', '0x00000060']:
#                             verify_res.append(True)
#                         else:
#                             verify_res.append(False)
#                     else:
#                         value_read.append(out[2:])
#                         if out.split()[-1] == self.check_items[command]:
#                             verify_res.append(True)
#                         else:
#                             verify_res.append(False)
#                 else:
#                     logging.info("Invalid command: " + command)
#     def verify_jesd(self, repeat):
#         """
#
#         :return:
#         """
#         results = []
#         thread_pool = []
#         for i in range(len(self.made_ips)):
#             t = threading.Thread(target=self._verify_jesd_single_ip, args=(self.made_ips[i], repeat, results, i))
#             thread_pool.append(t)
#             t.start()
#
#         for t in thread_pool:
#             t.join()
#
#         return results # a list of values read



class JesdVerify(object):

    def __init__(self, made_ips, product_type):
        self.ips = made_ips
        self.product_type = product_type
        self.check_items = self._load_json()
        self.fail_flag = False
        self.logger = logging.getLogger("main")

    def _load_json(self):
        if self.product_type == "AEQZ":
            json_file = "AEQZ_JESD_CHECK.json"
        elif self.product_type == "AEQV" or self.product_type == "AENB" or self.product_type == "AEQY":
            json_file = "AEQV_JESD_CHECK.json"
        elif self.product_type == "AEQE":
            json_file = "AEQE_JESD_CHECK.json"
        # TODO
        else:
            json_file = "AEQZ_JESD_CHECK.json"
        current_path = os.path.dirname(os.path.realpath(__file__))
        with open(os.path.join(current_path, "JESDTestItems", json_file), "r") as f:
            check_items = json.load(f)
        return check_items

    def _load_commands(self, row, col):
        self.commands = []
        if self.product_type == "AEQZ":
            command_file = "JESDCommands_AEQZ.txt"
        elif self.product_type == "AEQV" or self.product_type == "AENB" or self.product_type == "AEQY":
            command_file = "JESDCommands_AEQV.txt"
        # TODO
        else:
            command_file = "JESDCommands_AEQZ.txt"
        current_path = os.path.dirname(os.path.realpath(__file__))
        with open(os.path.join(current_path, 'JesdTestItems', command_file), 'r') as f:
            lines = f.readlines()
        for line in lines:
            if not line.startswith('#'):
                self.commands.append(line.strip())
                # command_to_write = re.search(r'/dev/spidev.*$', line)
                # if command_to_write:
                #     _ = self.ws.cell(column=col, row=row, value=command_to_write.group().strip())
                # else:
                #     _ = self.ws.cell(column=col, row=row, value=line.strip())
                # row += 1
        # self.wb.save(filename='JesdVerify.xls')

    def _read_values(self, row, col, scp, summary, ip):
        value_wrong = False
        self.ws.cell(row=row, column=col, value=time.strftime("%m%d%H%M%S"))
        row += 1
        for command in self.commands:
            command = command.strip()
            # logging.info(stderr.readline())
            # check with 0xfc90010c
            if command in self.check_items.keys():
                stdin, stdout, stderr = scp.write(command)
                out = stdout.readline().strip()
                # print("out:"+out)
                logging.debug("get respond => %s" % out)
                if not self.check_items[command]:
                    pass
                elif self.check_items[command] == "LAST2BIT":
                    logging.debug('get binary value => %s' % bin(int(out, 16)))
                    _ = self.ws.cell(row=row, column=col, value=bin(int(out, 16)))
                    if repr(bin(int(out, 16)))[-3:-1] == '11':
                        logging.debug('Verification result => correct')
                    else:
                        value_wrong = True
                        if command in summary[ip].keys():
                            summary[ip][command] += 1
                        else:
                            summary[ip][command] = 1
                        logging.debug('Verification result => wrong')
                        _.font = Font(color='00FF0000')
                elif self.check_items[command] == "CHECKLINK":
                    _ = self.ws.cell(row=row, column=col, value=out)
                    if out in ['0x00002860', '0x00000060', '0x00008060']:
                        logging.debug('Verification result => correct')
                    else:
                        value_wrong = True
                        if command in summary[ip].keys():
                            summary[ip][command] += 1
                        else:
                            summary[ip][command] = 1
                        logging.debug('Verification result => wrong')
                        _.font = Font(color='00FF0000')
                else:
                    _ = self.ws.cell(row=row, column=col, value=out[-2:])
                    if out.split()[-1] == self.check_items[command]:
                        logging.debug('Verification result => correct')
                    else:
                        value_wrong = True
                        if command in summary[ip].keys():
                            summary[ip][command] += 1
                        else:
                            summary[ip][command] = 1
                        logging.debug('Verification result => wrong')
                        _.font = Font(color='00FF0000')
                row += 1
            else:
                self.logger.info("Invalid command: " + command)
        return value_wrong

    def JesdTest_singleIP(self, REPEAT, ip, row, col, summary):
        pass_times = 0
        fail_times = 0

        scp = SCP(ip, 22)
        scp.connect_noPassword()
        # create ssh with no password
        current_path = os.path.dirname(os.path.realpath(__file__))
        local = os.path.join(current_path, 'JesdTestItems/spiafe')
        remote = '/var/tmp/spiafe'
        scp.upload(local, remote)
        # print("ip", ip, "upload done")

        for i in range(REPEAT):
            get_wrong_value = self._read_values(row, col, scp, summary, ip)
            if get_wrong_value:
                self.fail_flag = True
                fail_times += 1
            else:
                pass_times += 1
            col += 1
            self.logger.info(f"{ip} read finished...get fail values {get_wrong_value}")
            time.sleep(1)
        scp.close()
        if "pass time" in summary[ip].keys():
            summary[ip]["pass time"] += pass_times
        else:
            summary[ip]["pass time"] = pass_times

        if "fail time" in summary[ip].keys():
            summary[ip]["fail time"] += fail_times
        else:
            summary[ip]["fail time"] = fail_times

        # self.ws.cell(row=row, column=col-2, value=ip)
        # self.ws.cell(row=row, column=col-1, value="Read counts:" + str(REPEAT))
        # self.ws.cell(row=row, column=col, value="PASS: " + str(pass_times) + " FAIL: " + str(fail_times))
        # if not self.ws.cell(row=row, column=1).value:
        #     self.ws.cell(row=row, column=1, value="0/0")
        # total_pass = self.ws.cell(row=row, column=1).value.split("/")[0]
        # total_fail = self.ws.cell(row=row, column=1).value.split("/")[1]
        # total_pass = int(total_pass) + pass_times
        # total_fail = int(total_fail) + fail_times
        # self.ws.cell(row=row, column=1, value=str(total_pass)+"/"+str(total_fail))

    def JesdTest(self, REPEAT=1):
        current_path = os.path.dirname(os.path.realpath(__file__))
        try:
            self.wb = load_workbook(os.path.join(current_path, 'MainTestResults/MainResult.xlsx'))
        except FileNotFoundError:
            self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "JESD Raw data"
        row = 1
        col = self.ws.max_column + 1
        self._load_commands(row, col)

        summary_file = f"JesdLogAndResults/jesd.json"
        try:
            with open(os.path.join(current_path, summary_file), 'r') as f:
                summary = json.load(f)
        except FileNotFoundError:
            summary = {'summary': {"Cycle times": 0, "Pass times": 0, "Fail times": 0, "read counts": 0}}
            for ip in self.ips:
                summary[ip] = {'pass time': 0, 'fail time': 0}

        threads = []
        for ip in self.ips:
            # self.logger.info("IP: "+ip)
            t = threading.Thread(target=self.JesdTest_singleIP, args=(REPEAT, ip, row, col, summary))
            row += len(self.commands) + 2
            t.start()
            threads.append(t)
        for t in threads:
            t.join()
        # print("done")

        if self.fail_flag:
            download_runtime_log(path=os.path.join(current_path, 'MainTestResults/runtimeLog/'), log_name='Jesd_fail')
            summary['summary']["Fail times"] += 1
        else:
            summary['summary']["Pass times"] += 1
        summary['summary']["Cycle times"] += 1
        summary['summary']['read counts'] += REPEAT

        with open(os.path.join(current_path, 'JesdLogAndResults/jesd.json'), 'w') as f:
            json.dump(summary, f)

        self.wb.save(filename=os.path.join(current_path, 'MainTestResults/MainResult.xlsx'))
        return not self.fail_flag

    # def generate_summary(self):


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s:%(levelname)s:%(message)s")
    handler = JesdVerify(made_ips=['192.168.101.2', '192.168.101.3', '192.168.101.4', '192.168.101.5'],
                         product_type="AEQZ")
    handler.JesdTest(3)
