import os
import time
import logging
import re
from openpyxl import Workbook
from openpyxl.styles import Font
current_path = os.path.dirname(os.path.realpath(__file__))
# sys.path.insert(0, os.path.dirname(os.path.dirname(current_path)))
# add tools_dev into sys path
# not necessary if call from start.py
from tools_dev.connections.miniftp import SCP

class JesdVerify(object):
    def __init__(self):
        self.UL_JESD_status_commands = [
            '/sbin/devmem 0xfc980270',
            '/sbin/devmem 0xfc980274',
            '/sbin/devmem 0xfc980278',
            '/sbin/devmem 0xfc980294',
            '/sbin/devmem 0xfc98027C',
            '/sbin/devmem 0xfc980280',
            '/sbin/devmem 0xfc980288',
            '/sbin/devmem 0xfc98028C',
            '/sbin/devmem 0xfc980290',
            '/sbin/devmem 0xfc98026C',
            '/sbin/devmem 0xfc980264',
            '/sbin/devmem 0xfc980268']

        self.UL_int_status_commands = [
            '/sbin/devmem 0xfc980094',
            '/sbin/devmem 0xfc980098',
            '/sbin/devmem 0xfc98009c',
            '/sbin/devmem 0xfc9800a0',
            '/sbin/devmem 0xfc9800a4',
            '/sbin/devmem 0xfc9800a8',
            '/sbin/devmem 0xfc9800ac',
            '/sbin/devmem 0xfc9800b0',
            '/sbin/devmem 0xfc9800b8',
            '/sbin/devmem 0xfc9800bc',
            '/sbin/devmem 0xfc9800c0',
            '/sbin/devmem 0xfc9800c4']

        self.reg_values = {'0x12a': 'AA', '0x12c': '55'}
        self.pass_times = 0
        self.fail_times = 0
        self.wb = Workbook()
        self.ws = self.wb.active
        # self.ws.title = "JesdVerify"

    def _scp_connect(self):
        self.scp = SCP('192.168.101.2', 22)
        self.scp.connect_noPassword()
        # create ssh with no password
        local = os.path.join(current_path, 'spiafe')
        remote = '/var/tmp/spiafe'
        self.scp.upload(local, remote)

    def _load_commands(self):
        self.commands = []
        row = 2
        with open(os.path.join(current_path, "JESDCommands.txt"), 'r') as f:
            lines = f.readlines()
        for line in lines:
            if not line.startswith('#'):
                self.commands.append(line.strip())
                command_to_write = re.search(r'/dev/spidev.*$', line)
                if command_to_write:
                    _ = self.ws.cell(column=1, row=row, value=command_to_write.group().strip())
                else:
                    _ = self.ws.cell(column=1, row=row, value=line.strip())
                row += 1
        # self.wb.save(filename='JesdVerify.xls')

    def _scp_close(self):
        self.scp.close()

    def _read_values(self, col):
        row = 2
        for command in self.commands:
            # logging.info(f'run command {command}')
            stdin, stdout, stderr = self.scp.write(command)
            out = stdout.readline().strip()
            logging.info("get respond => %s" % out)
            command = command.strip()
            # logging.info(stderr.readline())
            # check with 0xfc90010c
            if command == '/sbin/devmem 0xfc90010c':
                _ = self.ws.cell(row=row, column=col, value=out)
                if out == '0xFFFFFFFF':
                    logging.info('Verification result => correct')
                else:
                    logging.info('Verification result => wrong')
                    self.test_result = 'FAIL'
                    _.font = Font(color='00FF0000')

            elif command in self.UL_JESD_status_commands:
                logging.info('get binary value => %s' % bin(int(out, 16)))
                _ = self.ws.cell(row=row, column=col, value=bin(int(out, 16)))
                if repr(bin(int(out, 16)))[-3:-1] == '11':
                    logging.info('Verification result => correct')
                else:
                    self.test_result = 'FAIL'
                    logging.info('Verification result => wrong')
                    _.font = Font(color='00FF0000')

            elif command in self.UL_int_status_commands:
                _ = self.ws.cell(row=row, column=col, value=out)
                if out in ['0x00002860', '0x00000060']:
                    logging.info('Verification result => correct')
                else:
                    self.test_result = "FAIL"
                    logging.info('Verification result => wrong')
                    _.font = Font(color='00FF0000')


            # write values
            # should return Register to write 0015 with value 02
            elif re.search(r'-w .* -v .*', command):
                _ = self.ws.cell(row=row, column=col, value=out)
                if out.startswith('Register to write'):
                    logging.info('write value successfully')
                else:
                    logging.info('write value falied')
                    self.test_result = 'FAIL'
                    _.font = Font(color='00FF0000')
            # read values
            #  should return Value from 012A is: AA
            elif re.search(r'-r .*', command):
                reg = re.search(r'-r .*', command).group()[-5:]
                _ = self.ws.cell(row=row, column=col, value=out[-2:])
                if out[-2:] == self.reg_values[reg]:
                    logging.info('Verification result => correct')
                else:
                    self.test_result = 'FAIL'
                    logging.info('Verification result => wrong')
                    _.font = Font(color='00FF0000')
            else:
                # logging.info("Invalid command!")
                _ = self.ws.cell(row=row, column=col, value=out)
                _.font = Font(color='00FF0000')
            row += 1

    def JesdTest(self, REPEAT):
        self.test_result = 'PASS'
        try:
            self._load_commands()
            self._scp_connect()
            col = 2
            for i in range(REPEAT):
                self._read_values(col)
                col += 1
                if self.test_result == 'PASS':
                    self.pass_times += 1
                else:
                    self.fail_times += 1
                time.sleep(3)

            logging.info('JESD TEST RESULT => ' + self.test_result)
            logging.info('#' * 100)

        except Exception as e:
            logging.info(msg=e)
            logging.info('error: ', e)

        self._scp_close()
        self.ws.cell(row=1, column=1, value="Test loops:" + str(REPEAT))
        self.ws.cell(row=1, column=2, value="PASS: " + str(self.pass_times) + " FAIL: " + str(self.fail_times))
        current_path = os.path.dirname(os.path.realpath(__file__))
        self.wb.save(filename=os.path.join(current_path, 'JesdLogAndResults/JesdVerify.xlsx'))


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, filename=os.path.join(current_path, 'JesdLogAndResults/JesdVerify.log'), format="%(asctime)s:%(levelname)s:%(message)s")
    REPEAT = 5
    handler = JesdVerify()
    handler.JesdTest(REPEAT)
