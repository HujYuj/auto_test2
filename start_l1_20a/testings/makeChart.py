import json
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from openpyxl.chart import (
    PieChart,
    BarChart,
    Reference
)
from openpyxl.chart.label import DataLabelList

current_path = os.path.dirname(os.path.realpath(__file__))

print(current_path)


def make_chart_carrier(result_file="MainTestResults/MainResult.xlsx", summary_file="CarrierTestResults"
                                                                                   "/CarrierTestResult.json"):
    try:
        wb = load_workbook(os.path.join(current_path, result_file))
    except:
        wb = Workbook()
    if "Carrier Summary" in wb.sheetnames:
        del wb["Carrier Summary"]

    wb.create_sheet("Carrier Summary")
    ws = wb["Carrier Summary"]
    _ = ws.cell(row=1, column=1, value='Carrier Summary')
    _.font = Font(bold=True)
    # summary_file = "CarrierTestResults/CarrierTestResult.json"
    with open(os.path.join(current_path, summary_file), 'r') as f:
        res = json.load(f)
    ips = [key for key in res.keys() if key != 'summary']
    row = 1
    pie_row_start = 4
    for ip in ips:
        data_to_add = [["made ip", ip]]
        for key in res[ip].keys():
            data_to_add.append([key, res[ip][key]])
        for line in data_to_add:
            ws.append(line)
        pie_row_end = pie_row_start + 3
        pie = PieChart()
        print("pie row start:", pie_row_start)
        print("pie row end:", pie_row_end)
        labels = Reference(ws, min_col=1, min_row=pie_row_start, max_row=pie_row_end)
        data = Reference(ws, min_col=2, min_row=pie_row_start - 1, max_row=pie_row_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = ip + " fail rate"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws.add_chart(pie, "D" + str(row))
        pie_row_start += len(data_to_add)
        row += 18

    wb.save(os.path.join(current_path, result_file))


def make_chart_jesd(result_file="MainTestResults/MainResult.xlsx", summary_file="JesdLogAndResults/jesd.json"):
    try:
        wb = load_workbook(os.path.join(current_path, result_file))
    except:
        wb = Workbook()
    if "JESD Summary" in wb.sheetnames:
        del wb["JESD Summary"]

    wb.create_sheet("JESD Summary")
    ws = wb["JESD Summary"]
    _ = ws.cell(row=1, column=1, value='JESD Summary')
    _.font = Font(bold=True)
    ws.cell(row=2, column=1, value="Cycle times")
    ws.cell(row=3, column=1, value="Pass times")
    ws.cell(row=4, column=1, value="Fail times")
    ws.cell(row=5, column=1, value="Read counts")

    with open(os.path.join(current_path, summary_file), 'r') as f:
        res = json.load(f)
        ws.cell(row=2, column=2, value=res['summary']["Cycle times"])
        ws.cell(row=3, column=2, value=res['summary']["Pass times"])
        ws.cell(row=4, column=2, value=res['summary']["Fail times"])
        ws.cell(row=5, column=2, value=res['summary']["read counts"])

    row = 7
    pie_row_start = 7
    bar_row_start = 9
    ips = [key for key in res.keys() if key != 'summary']
    for ip in ips:
        # wrong_value_file = f"JesdLogAndResults/wrong_value_record_{ip.split('.')[-1]}.json"
        # with open(os.path.join(current_path, wrong_value_file), 'r') as f:
        #     res = json.load(f)
        data_to_add = [["made ip", ip]]
        for key in res[ip].keys():
            reg = re.search(r'/sbin/devmem .*$', key)
            if reg:
                data_to_add.append([reg.group().split()[-1], res[ip][key]])
            else:
                data_to_add.append([key, res[ip][key]])
        bar_rows = len(data_to_add) - 3
        pie_row_end = pie_row_start + 1
        bar_row_end = bar_row_start + bar_rows - 1

        for line in data_to_add:
            ws.append(line)

        pie = PieChart()
        print("pie row start:", pie_row_start)
        print("pie row end:", pie_row_end)
        labels = Reference(ws, min_col=1, min_row=pie_row_start, max_row=pie_row_end)
        data = Reference(ws, min_col=2, min_row=pie_row_start - 1, max_row=pie_row_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = ip + " fail rate"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws.add_chart(pie, "D" + str(row))

        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = ip + " failed register"
        bar.y_axis.title = 'fail times'
        bar.x_axis.title = 'fail case'

        data = Reference(ws, min_col=2, min_row=bar_row_start - 1, max_row=bar_row_end)
        cats = Reference(ws, min_col=1, min_row=bar_row_start, max_row=bar_row_end)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.shape = 4
        ws.add_chart(bar, "N" + str(row))
        row += 18
        print(len(data))
        bar_row_start += len(data_to_add)
        pie_row_start += len(data_to_add)
    wb.save(os.path.join(current_path, result_file))


def make_chart_main(result_file="MainTestResults/MainResult.xlsx", result_json='MainTestResults/main_result.json'):
    try:
        wb = load_workbook(os.path.join(current_path, result_file))
        if "Main Summary" in wb.sheetnames:
            del wb['Main Summary']
        wb.create_sheet('Main Summary')
        ws = wb['Main Summary']
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Main Summary'

    with open(os.path.join(current_path, result_json), 'r') as f:
        res = json.load(f)

    for key in res.keys():
        ws.append([key, res[key]])

    pie_row_start = 2
    pie_row_end = pie_row_start + len(res) - 2
    pie = PieChart()
    print("pie row start:", pie_row_start)
    print("pie row end:", pie_row_end)
    labels = Reference(ws, min_col=1, min_row=pie_row_start, max_row=pie_row_end)
    data = Reference(ws, min_col=2, min_row=pie_row_start - 1, max_row=pie_row_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "main cycle fail rate"
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "D2")

    wb.save(os.path.join(current_path, result_file))


if __name__ == "__main__":
    make_chart_jesd("MainTestResults/MainResult_1110.xlsx", "JesdLogAndResults/jesd_1110.json")
    make_chart_main("MainTestResults/MainResult_1110.xlsx", 'MainTestResults/main_result_1110.json')
    make_chart_carrier("MainTestResults/MainResult_1110.xlsx", "CarrierTestResults/CarrierTestResult_1110.json")
